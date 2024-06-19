#
# @copyright
# Copyright (c) 2022 OVTeam
#
# All Rights Reserved
#
# Licensed under the MIT License;
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# https://choosealicense.com/licenses/mit/
#

from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook


class XLSX:
    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = None
        self.allow_save = False

    def set_dst_file_path(self, file_path):
        self.dst_file_path = file_path

    def load_file(self, is_template = True):
        self.wb = load_workbook(self.file_path, keep_links=False)
        self.wb.template = is_template

    def read(self, parserObj):
        wb = load_workbook(self.file_path, read_only=True,
                           data_only=True, keep_links=False)
        if self.sheet_name == None:
            if "Sheet1" in wb.sheetnames:
                self.sheet_name = "Sheet1"
            else:
                self.sheet_name = wb.sheetnames[0]

        ws = wb[self.sheet_name]
        idx = 0
        for row in ws.iter_rows(values_only=True):
            if idx == 0:
                parserObj.detect_header_map(row)
            else:
                parserObj.process(row)
                if idx == 1:
                    parserObj.validate()
            idx += 1
        parserObj.finish()

    def add_row(self, rows):
        if self.wb != None:
            self.allow_save = True
            ws = self.wb[self.sheet_name]
            ws.append(rows)

    def save(self, file_path = None):
        if self.allow_save:
            if file_path == None:
                file_path = self.dst_file_path
            if self.wb != None:
                self.wb.save(file_path)
                return True
        return False
